"""
Paso 1 output: annotate segments with AIM candidates and write the
parsimonious intermediate Excel/CSV consumed by humans and by Paso 2.

Main-sheet columns are EXACTLY the agreed contract (docs/intermediate_schema.md):

    case_id, system, segment_id, parent_segment_id, segment_order,
    title, chapter, article, segment_type, is_list_item, is_substantive,
    segment_text, aim_n_suggested, aim_text_candidate, aim_trigger_candidate,
    aim_candidate_confidence, has_own_aim_candidate, needs_review,
    review_reason, model_version, source_base

Nothing taxonomic is produced here (no TYPE/TAXON/LINK, no link-grammar /
link-aim suggestions, no normalized verbs). Technical metadata (character
offsets, lemmas) goes to a sidecar ``*_technical.json``, never to the Excel.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Optional

import pandas as pd

import re

from .aim_extractor import extract_aims
from .segmenter import Segment, segment_document
from .sentence_splitter import _load_model


def case_id_from_filename(path: str | Path) -> int | None:
    """Leading number of the filename, when present ("7. Reglamento..." -> 7)."""
    match = re.match(r"\s*(\d+)", Path(path).name)
    return int(match.group(1)) if match else None

PASO1_COLUMNS = [
    "case_id", "system", "segment_id", "parent_segment_id", "segment_order",
    "title", "chapter", "article", "segment_type", "is_list_item",
    "is_substantive", "segment_text", "aim_n_suggested", "aim_text_candidate",
    "aim_trigger_candidate", "aim_candidate_confidence",
    "has_own_aim_candidate", "needs_review", "review_reason",
    "model_version", "source_base",
]

# Mean BIO softmax confidence of CRF-decoded predictions sits around 0.93-0.97
# on real documents (median ~0.96, p20 ~0.87); 0.85 flags roughly the bottom
# fifth of candidates. Recalibrated 2026-06-10 on cases 20/23/24/25/13 after
# fixing the degenerate-tokenizer bug (the previous 0.55 was calibrated on
# [UNK]-sequence predictions and now flags almost nothing).
_LOW_CONFIDENCE = 0.85
_SHORT_SEGMENT_CHARS = 25

_SCHEMA_DESCRIPTIONS = {
    "case_id": "Numeric case identifier (from filename or caller).",
    "system": "Resource system (irrigation, fishery, ...), if known.",
    "segment_id": "Stable identifier of the segment within the document.",
    "parent_segment_id": "Segment this one structurally depends on (list header for items).",
    "segment_order": "Reading order in the document.",
    "title": "Current Titulo reference (structural context).",
    "chapter": "Current Capitulo/Chapter reference.",
    "article": "Current Articulo/Section reference.",
    "segment_type": "body_sentence | list_header | list_item | heading | preamble | signature | annex | metadata.",
    "is_list_item": "True when the segment is an item of a list.",
    "is_substantive": "Useful-context filter: True = candidate for coding; False = documentary context.",
    "segment_text": "Text of the segment, verbatim. Single text column (translate BEFORE Paso 1 if unsupported language).",
    "aim_n_suggested": "Model-suggested number of AIMs (pre-fill aid, NOT a final count).",
    "aim_text_candidate": "Candidate AIM fragment(s); multiple joined with ' | '.",
    "aim_trigger_candidate": "Verb/expression that appears to trigger each candidate; joined with ' | '.",
    "aim_candidate_confidence": "Model confidence for the candidates (0-1).",
    "has_own_aim_candidate": "True when the segment seems to carry its own AIM (neutral, no link typing).",
    "needs_review": "True when a human should look at this segment.",
    "review_reason": "Why review is suggested (';'-separated).",
    "model_version": "Model that produced the suggestions.",
    "source_base": "Source document file.",
}


def _first_verb(text: str, language: str) -> tuple[str, str]:
    """First VERB/AUX token in *text* (surface, lemma); empty strings if none."""
    if not text.strip():
        return "", ""
    doc = _load_model(language)(text)
    for token in doc:
        if token.pos_ in ("VERB", "AUX"):
            return token.text, token.lemma_.lower()
    return "", ""


def _candidate_fields(seg: Segment, aim_model, language: str) -> tuple[dict, dict]:
    """Model/heuristic AIM candidates for one segment.

    Returns (main_fields, technical) — technical carries offsets/lemmas for the
    sidecar JSON only.
    """
    spans: list[tuple[int, int, str]] = []
    confidence: Optional[float] = None
    if aim_model is not None:
        pred = aim_model.predict(seg.segment_text)
        spans = list(pred.spans)
        confidence = pred.confidence
        count = pred.count
    else:
        count = None

    segment_aims = None  # lazy: segment-level extraction for trigger fallback
    triggers: list[dict] = []
    for start, end, span_text in spans:
        aims = extract_aims(span_text, language=language)
        if not aims:
            # span fragment may lack parse context (e.g. "shall render his
            # report") — fall back to segment-level verbs inside the span
            if segment_aims is None:
                segment_aims = extract_aims(seg.segment_text, language=language)
            aims = [a for a in segment_aims if start <= a["position"] < end]
        if aims:
            triggers.append({"surface": aims[0]["verb"], "lemma": aims[0]["lemma"],
                             "start": start, "end": end, "span_text": span_text})
        else:
            # last resort: first verb/aux in the span, lexicon membership not
            # required ("expresion textual que parece gatillar el AIM")
            verb, lemma = _first_verb(span_text, language)
            triggers.append({"surface": verb, "lemma": lemma,
                             "start": start, "end": end, "span_text": span_text})

    if aim_model is None:
        # Heuristic fallback: institutional verbs found by the extractor
        aims = extract_aims(seg.segment_text, language=language)
        count = len(aims)
        triggers = [{"surface": a["verb"], "lemma": a["lemma"],
                     "start": a["position"], "end": None,
                     "span_text": seg.segment_text} for a in aims]
        spans = [(a["position"], -1, seg.segment_text) for a in aims]

    has_candidate = bool(count and count >= 1)
    main = {
        "aim_n_suggested": count,
        "aim_text_candidate": " | ".join(t["span_text"] for t in triggers) if triggers else "",
        "aim_trigger_candidate": " | ".join(t["surface"] for t in triggers if t["surface"]),
        "aim_candidate_confidence": round(confidence, 4) if confidence is not None else None,
        "has_own_aim_candidate": has_candidate,
    }
    technical = {"spans": triggers, "confidence": confidence}
    return main, technical


def _review_flags(seg: Segment, main: dict) -> tuple[bool, str]:
    reasons: list[str] = []
    text = seg.segment_text
    if seg.is_substantive and len(text) < _SHORT_SEGMENT_CHARS:
        reasons.append("short_segment")
    conf = main["aim_candidate_confidence"]
    if seg.is_substantive and main["has_own_aim_candidate"] and conf is not None and conf < _LOW_CONFIDENCE:
        reasons.append("low_confidence")
    if (seg.segment_type == "body_sentence" and text[:1].islower()):
        reasons.append("possible_bad_split")
    if seg.is_list_item and seg.parent_segment_id is None:
        reasons.append("list_item_no_parent")
    if main["aim_n_suggested"] is not None and main["aim_n_suggested"] >= 3:
        reasons.append("high_aim_count")
    if not seg.is_substantive and main["has_own_aim_candidate"]:
        reasons.append("context_filter_conflict")
    reasons.extend(seg.notes)
    return (bool(reasons), ";".join(reasons))


def pipeline_paso1(
    path: str | Path,
    aim_model=None,
    language: str = "es",
    case_id: int | None = None,
    system: str | None = None,
    model_version: str | None = None,
) -> tuple[pd.DataFrame, dict]:
    """Run Paso 1 on a document. Returns (DataFrame, technical sidecar dict)."""
    path = Path(path)
    if case_id is None:
        case_id = case_id_from_filename(path)
    if model_version is None:
        model_version = ("is-identifier-1.0" if aim_model is not None
                         else "heuristic_extractor")

    segments = segment_document(str(path), language=language)

    rows: list[dict] = []
    technical: dict[str, dict] = {}
    for seg in segments:
        main, tech = _candidate_fields(seg, aim_model, language)
        needs_review, review_reason = _review_flags(seg, main)
        rows.append({
            "case_id": case_id,
            "system": system,
            "segment_id": seg.segment_id,
            "parent_segment_id": seg.parent_segment_id,
            "segment_order": seg.segment_order,
            "title": seg.title,
            "chapter": seg.chapter,
            "article": seg.article,
            "segment_type": seg.segment_type,
            "is_list_item": seg.is_list_item,
            "is_substantive": seg.is_substantive,
            "segment_text": seg.segment_text,
            **main,
            "needs_review": needs_review,
            "review_reason": review_reason,
            "model_version": model_version,
            "source_base": path.name,
        })
        technical[seg.segment_id] = tech

    df = pd.DataFrame(rows, columns=PASO1_COLUMNS)
    return df, technical


# Column widths for the segments sheet (others get a small default).
_COL_WIDTHS = {
    "segment_id": 11, "parent_segment_id": 11, "segment_type": 14,
    "title": 16, "chapter": 16, "article": 9,
    "segment_text": 80, "aim_text_candidate": 60, "aim_trigger_candidate": 18,
    "review_reason": 24, "model_version": 18, "source_base": 22,
}
_REVIEW_FILL = "FFF3CD"  # soft amber for needs_review rows


def _summary_frame(df: pd.DataFrame) -> pd.DataFrame:
    """Counts a reviewer wants before opening the segments sheet."""
    rows: list[tuple[str, str, int]] = []
    rows.append(("document", "case_id", df["case_id"].iloc[0] if len(df) else None))
    rows.append(("document", "source_base", df["source_base"].iloc[0] if len(df) else ""))
    rows.append(("document", "model_version", df["model_version"].iloc[0] if len(df) else ""))
    rows.append(("document", "n_segments", len(df)))
    for val, n in df["segment_type"].value_counts().items():
        rows.append(("segment_type", str(val), int(n)))
    for val, n in df["is_substantive"].value_counts().items():
        rows.append(("is_substantive", str(val), int(n)))
    for val, n in df["needs_review"].value_counts().items():
        rows.append(("needs_review", str(val), int(n)))
    reasons = (df.loc[df["review_reason"] != "", "review_reason"]
               .str.split(";").explode().value_counts())
    for val, n in reasons.items():
        rows.append(("review_reason", str(val), int(n)))
    counts = df.loc[df["is_substantive"], "aim_n_suggested"].value_counts(dropna=False)
    for val, n in sorted(counts.items(), key=lambda kv: str(kv[0])):
        rows.append(("aim_n_suggested", str(val), int(n)))
    return pd.DataFrame(rows, columns=["section", "key", "value"])


def _format_segments_sheet(ws, df: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    n_cols = len(PASO1_COLUMNS)
    for idx, col in enumerate(PASO1_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = _COL_WIDTHS.get(col, 9)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    text_col = PASO1_COLUMNS.index("segment_text") + 1
    cand_col = PASO1_COLUMNS.index("aim_text_candidate") + 1
    wrap = Alignment(wrap_text=True, vertical="top")
    fill = PatternFill(start_color=_REVIEW_FILL, end_color=_REVIEW_FILL,
                       fill_type="solid")
    needs = df["needs_review"].tolist()
    for row_idx in range(2, len(df) + 2):
        ws.cell(row=row_idx, column=text_col).alignment = wrap
        ws.cell(row=row_idx, column=cand_col).alignment = wrap
        if needs[row_idx - 2]:
            for col_idx in range(1, n_cols + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    ws.auto_filter.ref = ws.dimensions
    # keep header row + id/order columns in view while scrolling
    ws.freeze_panes = ws.cell(row=2, column=PASO1_COLUMNS.index("title") + 1)


def write_paso1_excel(
    df: pd.DataFrame,
    out_path: str | Path,
    technical: dict | None = None,
) -> Path:
    """Write the Paso 1 Excel (segments + schema + summary) and technical sidecar.

    The segments sheet carries exactly PASO1_COLUMNS (contract order) with
    reviewer conveniences: auto-filter, frozen header/id columns, wrapped text
    and amber highlight on needs_review rows. The summary sheet aggregates
    counts; nothing taxonomic or administrative is added anywhere.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df[PASO1_COLUMNS].to_excel(writer, sheet_name="segments", index=False)
        schema_df = pd.DataFrame(
            [{"column": c, "description": _SCHEMA_DESCRIPTIONS[c]} for c in PASO1_COLUMNS]
        )
        schema_df.to_excel(writer, sheet_name="schema", index=False)
        _summary_frame(df).to_excel(writer, sheet_name="summary", index=False)

        _format_segments_sheet(writer.sheets["segments"], df)
        schema_ws = writer.sheets["schema"]
        schema_ws.column_dimensions["A"].width = 26
        schema_ws.column_dimensions["B"].width = 100
        summary_ws = writer.sheets["summary"]
        summary_ws.column_dimensions["A"].width = 18
        summary_ws.column_dimensions["B"].width = 34
        summary_ws.column_dimensions["C"].width = 40

    if technical is not None:
        tech_path = out_path.with_name(out_path.stem + "_technical.json")
        with open(tech_path, "w", encoding="utf-8") as f:
            json.dump(technical, f, ensure_ascii=False, indent=1, default=str)
    return out_path
