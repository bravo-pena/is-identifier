"""
Golden tests for the Paso 1 segmenter and parsimonious output schema.

Fragments mirror real patterns from the 14 markdown regulations:
inline lists ("Es responsabilidad de...: a) ... b) ..."), line-level bullets,
preambles, signatures, and article headings.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd

from is_identifier.paso1_output import PASO1_COLUMNS, pipeline_paso1, write_paso1_excel
from is_identifier.segmenter import Segment, _parse_block_lines, _split_inline_list, segment_document


# ---------------------------------------------------------------------------
# Inline list detection (collapsed "header: a) ... b) ..." lines)
# ---------------------------------------------------------------------------

def test_inline_list_header_and_items():
    text = ("Es responsabilidad de la asociación: a) la conservación de los canales; "
            "b) la información oportuna a los socios; c) convocar la asamblea general.")
    result = _split_inline_list(text)
    assert result is not None
    header, items = result
    assert header.endswith(":")
    assert "responsabilidad" in header
    assert len(items) == 3
    assert items[0].startswith("a)")
    assert "convocar" in items[2]


def test_inline_list_requires_colon_header():
    # Two markers but no colon header: too risky, do not split
    text = "El plazo es de 30 dias. 2. El pago se hara en efectivo."
    assert _split_inline_list(text) is None


# ---------------------------------------------------------------------------
# Line-level list parsing (markdown blocks keep newlines)
# ---------------------------------------------------------------------------

def test_parse_block_lines_items_and_continuation():
    block = ("Son obligaciones de los regantes:\n"
             "- Pagar la cuota anual.\n"
             "- Conservar las acequias\n"
             "en buen estado de limpieza.\n"
             "1. Asistir a las asambleas.")
    pieces = _parse_block_lines(block)
    kinds = [k for k, _ in pieces]
    assert kinds == ["text", "item", "item", "item"]
    # lowercase continuation line was appended to the second item
    assert "buen estado" in pieces[2][1]


# ---------------------------------------------------------------------------
# End-to-end segmentation of a small markdown document
# ---------------------------------------------------------------------------

_MD_DOC = """# Reglamento de la Comunidad

CONSIDERANDO que la Ley de Aguas establece el marco general.

## Artículo 1. De los fines

La Comunidad tiene por objeto el aprovechamiento de las aguas.

Son obligaciones de los regantes:

- Pagar la cuota anual que fije la Junta.
- Dignidad.
- Conservar las acequias en buen estado.

## Artículo 2. De las sanciones

El infractor deberá pagar una multa, de lo contrario será expulsado.

Firmado en Murcia, a 5 de mayo.
"""


def _write_doc(tmp_path: Path) -> Path:
    p = tmp_path / "test_reglamento_7.md"
    p.write_text(_MD_DOC, encoding="utf-8")
    return p


def test_segment_document_structure(tmp_path):
    doc = _write_doc(tmp_path)
    segs = segment_document(str(doc), language="es")
    by_type: dict[str, list[Segment]] = {}
    for s in segs:
        by_type.setdefault(s.segment_type, []).append(s)

    # Documentary context detected and NOT dropped
    assert any("CONSIDERANDO" in s.segment_text for s in by_type.get("preamble", []))
    assert any("Firmado" in s.segment_text for s in by_type.get("signature", []))
    for s in by_type.get("preamble", []) + by_type.get("signature", []):
        assert not s.is_substantive

    # Headings captured; article context propagated to body
    assert by_type.get("heading")
    body = by_type.get("body_sentence", [])
    assert any(s.article == "1" for s in body)

    # List preserved: header + 3 items, items linked to the header
    headers = by_type.get("list_header", [])
    items = by_type.get("list_item", [])
    assert len(items) == 3, [s.segment_text for s in items]
    header_ids = {h.segment_id for h in headers}
    assert all(i.parent_segment_id in header_ids for i in items)

    # The short bullet "Dignidad." survives as its own list_item (not junk)
    assert any("Dignidad" in i.segment_text for i in items)


def test_list_items_not_merged_into_parent(tmp_path):
    doc = _write_doc(tmp_path)
    segs = segment_document(str(doc), language="es")
    # No segment should contain both the header and an item text (old merging bug)
    for s in segs:
        if "Son obligaciones" in s.segment_text:
            assert "Pagar la cuota" not in s.segment_text


# ---------------------------------------------------------------------------
# Paso 1 pipeline contract (heuristic path, no model required)
# ---------------------------------------------------------------------------

def test_pipeline_paso1_contract_columns(tmp_path):
    doc = _write_doc(tmp_path)
    df, technical = pipeline_paso1(doc, aim_model=None, language="es")

    assert list(df.columns) == PASO1_COLUMNS
    # Paso 2 / administrative columns must never leak into the Paso 1 output
    # (explicit list from the approved contract, docs/intermediate_schema.md)
    forbidden_exact = {
        "TYPE", "TAXON", "LINK", "LINK.typ", "specificAIM",
        "suggested_link_role", "link_grammar_candidate", "link_aim_candidate",
        "verb_canonical", "verb_tense", "verb_mood", "verb_span_start",
        "verb_span_end", "verb_taxonomic_normalized",
        "Ini.Coder", "Rev.Coder", "Date.Ini.Code", "Date.Rev",
        "Version.Ini.Code", "Current.Version", "Reliability", "Notes",
        "Excerpt.Original.language", "Excerpt.Tranlsated",
    }
    forbidden = [c for c in df.columns
                 if c in forbidden_exact
                 or c.upper().startswith(("TYPE.", "TAXON.", "LINK."))
                 or "link_grammar" in c or "link_aim" in c]
    assert forbidden == []

    assert df["case_id"].iloc[0] is None or isinstance(df["case_id"].iloc[0], (int, float))
    assert (df["model_version"] == "heuristic_extractor").all()
    assert df["segment_text"].str.len().gt(0).all()
    # Short bullet flagged for review, not dropped
    dign = df[df["segment_text"].str.contains("Dignidad")]
    assert len(dign) == 1
    assert bool(dign["needs_review"].iloc[0])
    assert "short_segment" in dign["review_reason"].iloc[0]
    # technical sidecar has one entry per segment
    assert set(technical.keys()) == set(df["segment_id"])


def test_write_paso1_excel(tmp_path):
    doc = _write_doc(tmp_path)
    df, technical = pipeline_paso1(doc, aim_model=None, language="es")
    out = tmp_path / "paso1_out.xlsx"
    write_paso1_excel(df, out, technical)

    assert out.exists()
    back = pd.read_excel(out, sheet_name="segments")
    assert list(back.columns) == PASO1_COLUMNS
    schema = pd.read_excel(out, sheet_name="schema")
    assert set(schema["column"]) == set(PASO1_COLUMNS)
    assert out.with_name(out.stem + "_technical.json").exists()

    # summary sheet aggregates counts; reviewer formatting is applied
    summary = pd.read_excel(out, sheet_name="summary")
    assert list(summary.columns) == ["section", "key", "value"]
    assert "segment_type" in set(summary["section"])

    from openpyxl import load_workbook
    ws = load_workbook(out)["segments"]
    assert ws.auto_filter.ref is not None
    assert ws.freeze_panes is not None
    assert ws["A1"].font.bold
